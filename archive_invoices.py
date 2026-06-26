#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
发票归档脚本 v3
- 按记账期间（大文件夹）和凭证字号（小文件夹）自动归档发票 PDF/XML 文件
- 自动检测 Excel 表头列位置，适配不同系统导出格式
- 增量归档：复用已有文件夹，不删除已有文件
- 找不到发票时跳过且不创建空文件夹
- 修改下方默认路径或通过命令行传入

用法:
    python archive_invoices.py
    python archive_invoices.py --excel "D:\路径\xxx.xlsx" --dir "D:\发票目录"
"""

import os
import sys
import shutil
import openpyxl
import warnings
import re

warnings.filterwarnings('ignore')

# ── 默认配置：按需修改 ──
DEFAULT_EXCEL = r'D:\发票归档\20260520154557_.xlsx'
DEFAULT_DIR = r'D:\发票归档'

# 列名映射（自动从表头检测，关键词匹配）
COL_NAMES = {
    'invoice_num': '数电发票号码',
    'image': '发票影像',
    'voucher': '凭证字号',
    'period': '记账期间',
}


def detect_columns(ws, header_row=3):
    """从表头行自动检测列索引（1-based）"""
    header = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    col_map = {}
    for key, name in COL_NAMES.items():
        for i, h in enumerate(header, 1):
            if h and name in str(h):
                col_map[key] = i
                break
    print(f"自动检测列: {col_map}")
    for key in COL_NAMES:
        if key not in col_map:
            print(f"  ⚠️  未找到列: {key} ({COL_NAMES[key]})")
    return col_map


def get_folder_name(voucher_code):
    """
    凭证字号 → 文件夹名，省略前6位数字前缀
    例：202601 转-11 → 转-11
    """
    if not voucher_code or str(voucher_code).strip() in ('', ' '):
        return None
    code = str(voucher_code).strip()
    m = re.match(r'^\d{6}\s*(.*)', code)
    return m.group(1).strip() if m else code.strip()


def find_invoice_files(invoice_num, h_col, base_dir):
    """
    匹配发票 PDF 和 XML 文件（仅在 base_dir 根目录查找）
    优先 H 列精确匹配，备选 C 列发票号码模糊匹配
    """
    found = []

    # 1) H 列文件名精确匹配
    if h_col and str(h_col).strip():
        h_name = str(h_col).strip()
        pdf = os.path.join(base_dir, h_name)
        xml = os.path.join(base_dir, os.path.splitext(h_name)[0] + '.xml')
        if os.path.exists(pdf):
            found.append(pdf)
        if os.path.exists(xml):
            found.append(xml)
        if found:
            return found

    # 2) C 列发票号码模糊匹配
    if invoice_num and str(invoice_num).strip():
        num = str(invoice_num).strip()
        for fname in os.listdir(base_dir):
            fpath = os.path.join(base_dir, fname)
            if os.path.isfile(fpath) and num in fname:
                found.append(fpath)

    return found


def run(excel_path, invoice_dir):
    print(f"Excel: {excel_path}")
    print(f"发票目录: {invoice_dir}\n")

    wb = openpyxl.load_workbook(excel_path, read_only=False, data_only=True)
    ws = wb.active

    col_map = detect_columns(ws)
    if 'invoice_num' not in col_map or 'voucher' not in col_map:
        print("❌ 无法检测到必要列，退出")
        wb.close()
        return

    c_idx = col_map.get('invoice_num')
    h_idx = col_map.get('image', 8)
    k_idx = col_map['voucher']
    l_idx = col_map.get('period')

    records = []
    periods = set()

    for r in range(4, ws.max_row + 1):
        c_col = ws.cell(r, c_idx).value
        h_col = ws.cell(r, h_idx).value
        k_col = ws.cell(r, k_idx).value
        l_col = ws.cell(r, l_idx).value if l_idx else None

        if not c_col and not h_col:
            continue

        k_str = str(k_col).strip() if k_col else ''
        l_str = str(l_col).strip() if l_col else ''

        records.append({
            'row': r,
            'invoice_num': str(c_col).strip() if c_col else '',
            'h_col': str(h_col).strip() if h_col else '',
            'voucher_code': k_str,
            'period': l_str,
        })

        if l_str:
            periods.add(l_str)

    wb.close()

    print(f"共读取 {len(records)} 条发票记录")
    print(f"发现记账期间: {sorted(periods)}")

    no_voucher = [r for r in records if not r['voucher_code'] or r['voucher_code'] in ('', ' ')]
    if no_voucher:
        print(f"\n⚠️  无凭证字号的记录 ({len(no_voucher)} 条，将跳过):")
        for r in no_voucher:
            print(f"  行{r['row']}: {r['h_col'] or r['invoice_num']}")

    # 创建大文件夹
    print("\n=== 检查/创建记账期间文件夹 ===")
    for period in sorted(periods):
        pd = os.path.join(invoice_dir, period)
        if os.path.exists(pd):
            print(f"  已存在 {period}/")
        else:
            os.makedirs(pd, exist_ok=True)
            print(f"  ✓ 新建 {period}/")

    # 归档
    print("\n=== 开始归档发票 ===")
    moved = 0
    skipped = 0
    not_found = 0
    already = 0

    for rec in records:
        if not rec['voucher_code'] or rec['voucher_code'] in ('', ' '):
            skipped += 1
            continue
        if not rec['period']:
            print(f"  ⏭️  跳过（无记账期间）: {rec['h_col'] or rec['invoice_num']}")
            skipped += 1
            continue

        folder = get_folder_name(rec['voucher_code'])
        if not folder:
            skipped += 1
            continue

        files = find_invoice_files(rec['invoice_num'], rec['h_col'], invoice_dir)

        if not files:
            target_dir = os.path.join(invoice_dir, rec['period'], folder)
            h_name = rec['h_col']
            if h_name and os.path.exists(target_dir) and os.path.exists(os.path.join(target_dir, h_name)):
                already += 1
            else:
                print(f"  ❌ 未找到: {rec['h_col'] or rec['invoice_num']}（不创建文件夹）")
                not_found += 1
            continue

        target_dir = os.path.join(invoice_dir, rec['period'], folder)
        os.makedirs(target_dir, exist_ok=True)

        for fpath in files:
            fname = os.path.basename(fpath)
            dest = os.path.join(target_dir, fname)
            if os.path.exists(dest):
                print(f"  ⚠️  目标已存在，跳过: {fname}")
                already += 1
            else:
                shutil.move(fpath, dest)
                print(f"  ✓ {fname}  → {rec['period']}/{folder}/")
                moved += 1

    print(f"\n=== 归档完成 ===")
    print(f"  成功移动文件:   {moved} 个")
    print(f"  已在目标位置:   {already} 个（无需移动）")
    print(f"  跳过记录:       {skipped} 条")
    print(f"  文件未找到:     {not_found} 条")

    # 目录结构
    print("\n=== 最终目录结构 ===")
    for period in sorted(periods):
        pd = os.path.join(invoice_dir, period)
        if os.path.exists(pd):
            subdirs = sorted(
                [s for s in os.listdir(pd) if os.path.isdir(os.path.join(pd, s))],
                key=lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else 0
            )
            total = sum(
                sum(1 for f in os.listdir(os.path.join(pd, s)) if f.endswith('.pdf'))
                for s in subdirs
            )
            print(f"\n📁 {period}/  (共 {total} 张发票)")
            for sub in subdirs:
                sp = os.path.join(pd, sub)
                n_pdf = sum(1 for f in os.listdir(sp) if f.endswith('.pdf'))
                print(f"   📂 {sub}/  ({n_pdf} 张)")


if __name__ == '__main__':
    excel = DEFAULT_EXCEL
    invoice = DEFAULT_DIR
    args = sys.argv[1:]
    for i, a in enumerate(args):
        if a in ('--excel', '-e') and i + 1 < len(args):
            excel = args[i + 1]
        if a in ('--dir', '-d') and i + 1 < len(args):
            invoice = args[i + 1]
    run(excel, invoice)
